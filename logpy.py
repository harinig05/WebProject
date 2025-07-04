import os
import logging
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
folder_path = "C:/Users/gunas/Downloads/logexcel"
excel_path = os.path.join(folder_path, "test_result.xlsx")
log_file_path = os.path.join(folder_path, "test_log.log")
screenshot_path = os.path.join(folder_path, "products_page.png")
os.makedirs(folder_path, exist_ok=True)
logging.basicConfig(filename=log_file_path,level=logging.INFO,format='%(asctime)s - %(levelname)s - %(message)s')
logging.info("Test started")
try:
    driver = webdriver.Chrome()
    driver.get("https://www.saucedemo.com/")
    driver.maximize_window()
    logging.info("Opened SauceDemo site")

    driver.find_element(By.ID, "user-name").send_keys("standard_user")
    driver.find_element(By.ID, "password").send_keys("secret_sauce")
    driver.find_element(By.ID, "login-button").click()
    logging.info("Logged in")

    title_element = driver.find_element(By.CLASS_NAME, 'title')
    assert title_element.text == "Products", "Login failed"
    logging.info("Assertion passed — on Products page")

    driver.save_screenshot(screenshot_path)
    logging.info(f"Screenshot saved at {screenshot_path}")

    time.sleep(2)
    driver.quit()
    logging.info("Browser closed")

except Exception as e:
    logging.error(f"Test failed with exception: {e}")
    if 'driver' in locals():
        driver.quit()
try:
    workbook = load_workbook(excel_path)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Logs"
if sheet.max_row == 1 and sheet['A1'].value is None:
    sheet.append(["Timestamp", "Log Level", "Message"])
with open(log_file_path, "r") as log_file:
    for line in log_file:
        parts = line.strip().split(" - ", 2)
        if len(parts) == 3:
            sheet.append(parts)
workbook.save(excel_path)
print(" All logs written to Excel at:", excel_path)
